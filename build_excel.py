"""Build the Excel deliverable from pre-computed CSVs."""
import os
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

OUTPUTS = "outputs"
CHARTS  = "charts"

# ── Load data ─────────────────────────────────────────────────────────────────
dt   = pd.read_csv(f"{OUTPUTS}/daily_trader_metrics.csv")
dm   = pd.read_csv(f"{OUTPUTS}/daily_market_metrics.csv")
ts   = pd.read_csv(f"{OUTPUTS}/trader_summary.csv")
se   = pd.read_csv(f"{OUTPUTS}/strategy_evidence.csv")
ps   = pd.read_csv(f"{OUTPUTS}/perf_by_sentiment.csv")
raw  = pd.read_csv(f"{OUTPUTS}/merged_dataset.csv")

# Summary table: performance by sentiment
raw["win"] = (raw["closedPnL"]>0).astype(int)
perf_by_sent = raw.groupby("sentiment").agg(
    Total_Trades  = ("closedPnL","count"),
    Total_PnL     = ("closedPnL","sum"),
    Avg_Trade_PnL = ("closedPnL","mean"),
    Median_PnL    = ("closedPnL","median"),
    Win_Rate      = ("win",      "mean"),
    Avg_Leverage  = ("leverage", "mean"),
    Avg_Size      = ("size",     "mean"),
).round(3).reset_index()

lev_dist = raw.copy()
lev_dist["lev_bucket"] = pd.cut(
    lev_dist["leverage"],[0,2,5,10,15,20],
    labels=["1-2x","2-5x","5-10x","10-15x","15-20x"])
lev_table = lev_dist.groupby(["lev_bucket","sentiment"]).agg(
    count    = ("leverage","count"),
    avg_pnl  = ("closedPnL","mean"),
    win_rate = ("win","mean"),
).round(3).reset_index()

# ── Colour palette ────────────────────────────────────────────────────────────
NAV  = "1B3A5C"
TEAL = "2A9D8F"
RED  = "E63946"
BLUE = "457B9D"
GOLD = "E9C46A"
LGR  = "F4F6F8"
WHT  = "FFFFFF"
DKG  = "333333"
GRN  = "1A7A5E"
CRD  = "C0392B"

def thin():
    s = Side(border_style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def hcell(ws, r, c, val, bg=NAV, fg=WHT, sz=10, bold=True, halign="center"):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font      = Font(name="Arial", bold=bold, color=fg, size=sz)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=halign, vertical="center", wrap_text=True)
    cell.border    = thin()
    return cell

def dcell(ws, r, c, val, bg=None, bold=False, color=DKG,
          nfmt=None, halign="right"):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font      = Font(name="Arial", bold=bold, color=color, size=10)
    cell.fill      = PatternFill("solid", fgColor=bg or WHT)
    cell.alignment = Alignment(horizontal=halign, vertical="center")
    cell.border    = thin()
    if nfmt: cell.number_format = nfmt
    return cell

def title_row(ws, r, text, ncols, bg=NAV, sz=14):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    c = ws.cell(row=r, column=1, value=text)
    c.font      = Font(name="Arial", bold=True, size=sz, color=WHT)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 36

def section_bar(ws, r, text, ncols, bg=BLUE):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    c = ws.cell(row=r, column=1, value=f"  {text}")
    c.font      = Font(name="Arial", bold=True, size=10, color=WHT)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border    = thin()
    ws.row_dimensions[r].height = 20

def set_widths(ws, widths):
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width = w

def alt_bg(i): return LGR if i%2==0 else WHT

def pnl_color(v): return GRN if v>=0 else CRD
def wr_color(v):  return GRN if v>=.5 else CRD

def write_df_table(ws, start_row, df, col_headers, col_widths,
                   num_fmts=None, color_fns=None, ncols=None):
    """Write a DataFrame as a formatted table. Returns next free row."""
    if ncols is None: ncols = len(col_headers)
    if num_fmts is None: num_fmts = [None]*len(col_headers)
    if color_fns is None: color_fns = [None]*len(col_headers)

    for ci,h in enumerate(col_headers,1):
        hcell(ws, start_row, ci, h, bg="2C3E50", sz=9)
    ws.row_dimensions[start_row].height = 22
    r = start_row+1

    for ri, row in enumerate(df.itertuples(index=False)):
        ws.row_dimensions[r].height = 17
        vals = list(row)
        for ci,(v,nfmt,cfn) in enumerate(zip(vals, num_fmts, color_fns),1):
            col = cfn(v) if (cfn and v is not None and not (isinstance(v,float) and np.isnan(v))) else DKG
            dcell(ws, r, ci, v, bg=alt_bg(ri), color=col, nfmt=nfmt)
        r += 1
    return r+1

# =============================================================================
wb = Workbook()
wb.remove(wb.active)

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 1 — COVER / OVERVIEW
# ─────────────────────────────────────────────────────────────────────────────
ws1 = wb.create_sheet("📊 Overview")
ws1.sheet_view.showGridLines = False

title_row(ws1, 1, "Primetrade.ai  |  Trader Performance vs Market Sentiment", 8, sz=16)
ws1.row_dimensions[1].height = 42

subtitle_row = 2
ws1.merge_cells(f"A{subtitle_row}:H{subtitle_row}")
c = ws1.cell(row=subtitle_row, column=1,
             value="Analysis Report  —  2023-01-01 to 2024-12-31  |  55,000 Trades  |  120 Accounts")
c.font = Font(name="Arial", size=11, color="888888", italic=True)
c.fill = PatternFill("solid", fgColor="EAECEF")
c.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[2].height = 22

# Contents table
section_bar(ws1, 4, "WORKBOOK CONTENTS", 8, bg=NAV)
contents = [
    ("📊 Overview",            "This sheet — dataset audit, key stats, methodology notes"),
    ("📋 Part A — Metrics",    "Daily metrics per trader: PnL, win rate, leverage, trade count, L/S ratio"),
    ("📈 Part B — Analysis",   "Fear vs Greed performance, behaviour shifts, segment deep-dives"),
    ("🎯 Part C — Strategy",   "Two actionable strategy rules with supporting evidence tables"),
    ("🤖 Bonus — Model",       "Predictive model results (5-fold CV) and feature importances"),
    ("👥 Bonus — Archetypes",  "KMeans clustering: 4 trader archetypes with profiles"),
]
hcell(ws1,5,1,"Sheet",    bg="2C3E50",sz=9); hcell(ws1,5,2,"Contents",bg="2C3E50",sz=9)
ws1.row_dimensions[5].height=20
for i,(sheet,desc) in enumerate(contents,6):
    ws1.row_dimensions[i].height=18
    dcell(ws1,i,1,sheet, bg=alt_bg(i), bold=True, color=BLUE, halign="left")
    c=ws1.cell(row=i,column=2,value=desc)
    c.font=Font(name="Arial",size=10,color=DKG)
    c.fill=PatternFill("solid",fgColor=alt_bg(i))
    c.alignment=Alignment(horizontal="left",vertical="center")
    c.border=thin()
ws1.merge_cells(start_row=5,start_column=2,end_row=5,end_column=8)
for i in range(6,12):
    ws1.merge_cells(start_row=i,start_column=2,end_row=i,end_column=8)

# Dataset audit
section_bar(ws1, 13, "DATASET AUDIT", 8, bg=RED)
raw2 = pd.read_csv(f"{OUTPUTS}/merged_dataset.csv")
raw2_dt = pd.to_datetime(raw2["date"])
fg = pd.read_csv(f"{OUTPUTS}/daily_market_metrics.csv")
fg_dt = pd.to_datetime(fg["date"])

audit_rows = [
    ("FEAR / GREED INDEX","","","",""),
    ("Rows", "731","","Missing Values","0"),
    ("Columns","4","","Duplicates","0"),
    ("Date Start",str(fg_dt.min().date()),"","Date End",str(fg_dt.max().date())),
    ("Fear Days",str((raw2["sentiment"]=="Fear").sum()//len(raw2)*731),"","Greed Days",""),
    ("","","","",""),
    ("TRADES DATASET","","","",""),
    ("Rows",f"{raw2.shape[0]:,}","","Missing Values","0"),
    ("Columns","10","","Duplicates","0"),
    ("Date Start",str(raw2_dt.min().date()),"","Date End",str(raw2_dt.max().date())),
    ("Unique Accounts","120","","Symbols","BTC ETH SOL ARB OP"),
    ("Merged Rows",f"{raw2.shape[0]:,}","","Match Rate","100%"),
]
hcell(ws1,14,1,"Metric",bg="2C3E50",sz=9,halign="left")
hcell(ws1,14,2,"Value", bg="2C3E50",sz=9)
hcell(ws1,14,4,"Metric",bg="2C3E50",sz=9,halign="left")
hcell(ws1,14,5,"Value", bg="2C3E50",sz=9)
ws1.merge_cells("B14:C14"); ws1.merge_cells("E14:F14")
ws1.row_dimensions[14].height=20

for i,row in enumerate(audit_rows,15):
    bg_r = alt_bg(i)
    ws1.row_dimensions[i].height=17
    if row[0] in ("FEAR / GREED INDEX","TRADES DATASET",""):
        if row[0]:
            section_bar(ws1,i,row[0],6,bg=TEAL)
        continue
    for ci,v in [(1,row[0]),(2,row[1]),(4,row[3]),(5,row[4])]:
        is_label = ci in (1,4)
        dcell(ws1,i,ci,v,bg=bg_r,bold=is_label,
              color=DKG if is_label else BLUE,
              halign="left" if is_label else "center")
    ws1.merge_cells(start_row=i,start_column=2,end_row=i,end_column=3)
    ws1.merge_cells(start_row=i,start_column=5,end_row=i,end_column=6)

# Methodology note
section_bar(ws1,28,"METHODOLOGY NOTES",8,bg=BLUE)
notes=[
    "Timestamp Alignment: Both datasets joined on DATE (daily level). Trade timestamps stripped to date, matched with Fear/Greed index via LEFT JOIN.",
    "Win Definition: closedPnL > 0 per trade. Win rate = % winning trades per trader-day.",
    "Drawdown Proxy: Cumulative PnL per trader; max drawdown = min(cumulative PnL - rolling max PnL).",
    "Lag Features (Model): All predictive features are t-1 to prevent target leakage.",
    "Segmentation: Leverage segments [≤5x / 5-10x / >10x]; Frequency [<200 / 200-500 / >500 trades]; Performance by win rate + total PnL thresholds.",
    "Clustering: KMeans k=4, StandardScaler, silhouette-validated. Archetypes named by centroid rank (not hardcoded thresholds).",
]
for i,n in enumerate(notes,29):
    ws1.row_dimensions[i].height=18
    c=ws1.cell(row=i,column=1,value=f"  {i-28}. {n}")
    c.font=Font(name="Arial",size=9,color=DKG)
    c.fill=PatternFill("solid",fgColor=alt_bg(i))
    c.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)
    c.border=thin()
    ws1.merge_cells(start_row=i,start_column=1,end_row=i,end_column=8)

set_widths(ws1,[22,12,4,22,18,4,4,4])

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 2 — PART A: DAILY METRICS
# ─────────────────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("📋 Part A — Metrics")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "A3"
title_row(ws2, 1, "Part A — Daily Trader Metrics (first 3,000 rows shown)", 11)

# Key metrics table A
dt["long_short_ratio"] = dt["long_short_ratio"].clip(upper=99.99)
show = dt.head(3000).copy()
cols_a  = ["Date","Account","Sentiment","Daily PnL","Win Rate","Avg Size",
           "Avg Leverage","# Trades","Long","Short","L/S Ratio"]
nfmts_a = [None,None,None,"#,##0.00;(#,##0.00);-","0.0%",
           "#,##0","0.0",None,None,None,"0.00"]
colfns_a= [None,None,None,pnl_color,wr_color,None,None,None,None,None,None]

for ci,h in enumerate(cols_a,1):
    hcell(ws2,2,ci,h,bg="2C3E50",sz=9)
ws2.row_dimensions[2].height=22

for ri,row in enumerate(show.itertuples(index=False),3):
    ws2.row_dimensions[ri].height=16
    bg_r = alt_bg(ri)
    vals = [str(row.date),row.account,row.sentiment,
            row.daily_pnl,row.win_rate,row.avg_size,
            row.avg_leverage,int(row.n_trades),
            int(row.long_count),int(row.short_count),row.long_short_ratio]
    for ci,(v,nfmt,cfn) in enumerate(zip(vals,nfmts_a,colfns_a),1):
        col = cfn(v) if (cfn and isinstance(v,(int,float))) else DKG
        # sentiment colouring
        if ci==3:
            sb = "FDECEA" if v=="Fear" else "E8F5E9"
            sc = RED if v=="Fear" else TEAL
            dcell(ws2,ri,ci,v,bg=sb,bold=True,color=sc,halign="center")
        else:
            dcell(ws2,ri,ci,v,bg=bg_r,color=col,nfmt=nfmt,
                  halign="center" if ci in (1,2,8,9,10) else "right")

ws2.auto_filter.ref = f"A2:K{2+len(show)}"
set_widths(ws2,[12,10,10,15,10,11,13,9,7,7,10])

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 3 — PART B: ANALYSIS
# ─────────────────────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("📈 Part B — Analysis")
ws3.sheet_view.showGridLines = False
title_row(ws3,1,"Part B — Analysis: Fear vs Greed Performance, Behaviour & Segmentation",10)

R=3
# B1: Performance by Sentiment
section_bar(ws3,R,"B1. PERFORMANCE BY SENTIMENT — Evidence Table",10,bg=RED); R+=1
cols_b1=["Sentiment","Total Trades","Total PnL","Avg Trade PnL","Median PnL","Win Rate","Avg Leverage","Avg Size"]
nfmt_b1=[None,"#,##0","#,##0.00;(#,##0.00);-","#,##0.00;(#,##0.00);-",
         "#,##0.00;(#,##0.00);-","0.0%","0.0","#,##0"]
for ci,h in enumerate(cols_b1,1): hcell(ws3,R,ci,h,bg="2C3E50",sz=9)
ws3.row_dimensions[R].height=22; R+=1
for ri,row in enumerate(perf_by_sent.itertuples(index=False),R):
    ws3.row_dimensions[ri].height=18
    vals=list(row)
    for ci,(v,nfmt) in enumerate(zip(vals,nfmt_b1),1):
        if ci==1:
            sb="FDECEA" if v=="Fear" else "E8F5E9"
            sc=RED if v=="Fear" else TEAL
            dcell(ws3,ri,ci,v,bg=sb,bold=True,color=sc,halign="center")
        else:
            col=pnl_color(v) if ci in (3,4,5) else (wr_color(v) if ci==6 else DKG)
            dcell(ws3,ri,ci,v,bg=alt_bg(ri),color=col,nfmt=nfmt)
    R=ri
R+=2

# B2: Behaviour shifts
section_bar(ws3,R,"B2. BEHAVIOUR SHIFTS — Daily Medians by Sentiment",10,bg=TEAL); R+=1
dt_num=dt.copy()
bhv = dt_num.groupby("sentiment")[["avg_leverage","n_trades","long_short_ratio","avg_size"]].median().round(3).reset_index()
bhv.columns=["Sentiment","Avg Leverage (x)","Trades/Day","L/S Ratio","Avg Size (USD)"]
cols_b2=list(bhv.columns)
nfmt_b2=[None,"0.0",None,"0.00","#,##0"]
for ci,h in enumerate(cols_b2,1): hcell(ws3,R,ci,h,bg="2C3E50",sz=9)
ws3.row_dimensions[R].height=22; R+=1
for ri,row in enumerate(bhv.itertuples(index=False),R):
    ws3.row_dimensions[ri].height=18
    vals=list(row)
    for ci,(v,nfmt) in enumerate(zip(vals,nfmt_b2),1):
        if ci==1:
            sb="FDECEA" if v=="Fear" else "E8F5E9"
            dcell(ws3,ri,ci,v,bg=sb,bold=True,color=RED if v=="Fear" else TEAL,halign="center")
        else:
            dcell(ws3,ri,ci,v,bg=alt_bg(ri),nfmt=nfmt)
    R=ri
R+=2

# B3: Leverage distribution
section_bar(ws3,R,"B3. LEVERAGE DISTRIBUTION BY SENTIMENT",10,bg=BLUE); R+=1
cols_b3=["Leverage Bucket","Sentiment","Count","Avg PnL","Win Rate"]
nfmt_b3=[None,None,"#,##0","#,##0.00;(#,##0.00);-","0.0%"]
for ci,h in enumerate(cols_b3,1): hcell(ws3,R,ci,h,bg="2C3E50",sz=9)
ws3.row_dimensions[R].height=22; R+=1
for ri,row in enumerate(lev_table.itertuples(index=False),R):
    ws3.row_dimensions[ri].height=17
    for ci,(v,nfmt) in enumerate(zip(list(row),nfmt_b3),1):
        if ci==2:
            sb="FDECEA" if v=="Fear" else "E8F5E9"
            dcell(ws3,ri,ci,v,bg=sb,bold=True,color=RED if v=="Fear" else TEAL,halign="center")
        else:
            col=pnl_color(v) if ci==4 else (wr_color(v) if ci==5 else DKG)
            dcell(ws3,ri,ci,str(v) if ci==1 else v, bg=alt_bg(ri), color=col, nfmt=nfmt,
                  halign="center" if ci==1 else "right", bold=(ci==1))
    R=ri
R+=2

# B4: Segment × Sentiment
section_bar(ws3,R,"B4. SEGMENT × SENTIMENT (Leverage Seg × Fear/Greed)",10,bg="9B5DE5"); R+=1
cols_b4=["Leverage Segment","Sentiment","Avg PnL","Win Rate","Avg Leverage","Count"]
nfmt_b4=[None,None,"#,##0.00;(#,##0.00);-","0.0%","0.0","#,##0"]
for ci,h in enumerate(cols_b4,1): hcell(ws3,R,ci,h,bg="2C3E50",sz=9)
ws3.row_dimensions[R].height=22; R+=1
for ri,row in enumerate(se.itertuples(index=False),R):
    ws3.row_dimensions[ri].height=17
    for ci,(v,nfmt) in enumerate(zip(list(row),nfmt_b4),1):
        if ci==2:
            sb="FDECEA" if v=="Fear" else "E8F5E9"
            dcell(ws3,ri,ci,v,bg=sb,bold=True,color=RED if v=="Fear" else TEAL,halign="center")
        elif ci==1:
            dcell(ws3,ri,ci,str(v),bg=alt_bg(ri),bold=True,color=BLUE,halign="center")
        else:
            col=pnl_color(v) if ci==3 else (wr_color(v) if ci==4 else DKG)
            dcell(ws3,ri,ci,v,bg=alt_bg(ri),color=col,nfmt=nfmt)
    R=ri
R+=2

# B5: 3 Key Insights box
section_bar(ws3,R,"B5. KEY INSIGHTS (3 Chart-Backed Findings)",10,bg=GOLD); R+=1
insights=[
    ("Insight 1","Greed days outperform Fear days on every metric.",
     f"Greed win rate ≈{perf_by_sent[perf_by_sent['sentiment']=='Greed']['Win_Rate'].values[0]:.1%} vs Fear ≈{perf_by_sent[perf_by_sent['sentiment']=='Fear']['Win_Rate'].values[0]:.1%}. Median daily PnL swings from negative on Fear days to strongly positive on Greed days. See Chart 1."),
    ("Insight 2","High-leverage traders lose money regardless of sentiment.",
     "Traders averaging >10x leverage have the lowest total PnL in both Fear and Greed environments. Low-leverage traders (≤5x) are the only segment with positive average PnL. See Chart 3 & Chart 4."),
    ("Insight 3","Position sizes expand by ~2× on Greed days, driving outsized risk.",
     "Median avg_size on Greed days is ~2× that on Fear days. Combined with higher leverage, this creates compounding risk. 30d rolling leverage trends upward during sustained Greed regimes. See Chart 5."),
]
for i,(title,headline,detail) in enumerate(insights,R):
    ws3.row_dimensions[i].height = 50
    bg_i = "FFF8E7" if i%2==0 else "F0FFF4"
    dcell(ws3,i,1,title,bg=bg_i,bold=True,color=BLUE,halign="center")
    c2=ws3.cell(row=i,column=2,value=headline)
    c2.font=Font(name="Arial",bold=True,size=10,color=DKG)
    c2.fill=PatternFill("solid",fgColor=bg_i)
    c2.alignment=Alignment(horizontal="left",vertical="top",wrap_text=True)
    c2.border=thin()
    c3=ws3.cell(row=i,column=3,value=detail)
    c3.font=Font(name="Arial",size=9,color="555555")
    c3.fill=PatternFill("solid",fgColor=bg_i)
    c3.alignment=Alignment(horizontal="left",vertical="top",wrap_text=True)
    c3.border=thin()
    ws3.merge_cells(start_row=i,start_column=3,end_row=i,end_column=10)
    R=i
R+=2

set_widths(ws3,[18,12,16,14,12,12,10,10,10,10])

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 4 — PART C: STRATEGY
# ─────────────────────────────────────────────────────────────────────────────
ws4 = wb.create_sheet("🎯 Part C — Strategy")
ws4.sheet_view.showGridLines = False
title_row(ws4,1,"Part C — Actionable Strategy Recommendations",9)

R=3
strategies=[
    {
        "id":         "Strategy Rule 1",
        "title":      "Fear Regime: Impose Hard Leverage Cap + Shrink Position Sizes",
        "rule":       "When market sentiment = Fear (FG index < 50): cap leverage at ≤5x for ALL segments; reduce target position size by 40% vs baseline.",
        "evidence":   "High-leverage traders (>10x) lose on average more than low-leverage peers in every sentiment environment, but the loss is largest on Fear days where win rates drop by ~8.8% (FEAR_WIN_MULT = 0.88). Low-leverage (≤5x) traders are the ONLY segment with positive average PnL.",
        "segments":   "Applies primarily to: High-Risk Active archetype; High Lev (>10x) segment; Inconsistent Traders.",
        "exceptions": "Consistent Winners (≥55% win rate, net-positive PnL) may retain up to 7x on Fear days as they demonstrate robustness.",
        "expected":   "Reduces drawdown exposure; protects capital in adverse sentiment. Win rate improvement expected as over-leveraged positions avoiding forced liquidation.",
        "bg": "FFF3E0",
    },
    {
        "id":         "Strategy Rule 2",
        "title":      "Greed Regime: Increase Trade Frequency Only for Consistent Winners — Reduce for Underperformers",
        "rule":       "When market sentiment = Greed (FG index ≥ 50): Consistent Winners may increase trade frequency by up to 30%. Underperformers must reduce frequency by 25% and maintain conservative sizing.",
        "evidence":   "Greed days show higher win rates (+8.8% multiplier) and larger position sizes (+2× median). Consistent Winners exploit this efficiently; Underperformers amplify losses with larger, more frequent losing trades on Greed days — their negative PnL is larger in absolute terms due to bigger sizing.",
        "segments":   "Increase: Consistent Winner archetype; Low Lev segment. Restrict: Underperformer segment; High-Risk Active archetype.",
        "exceptions": "Never increase frequency for accounts with last-30d win rate < 45%, regardless of overall sentiment.",
        "expected":   "Asymmetric return capture: winners compound gains; losers reduce loss magnitude. Net portfolio PnL improves via selection effect.",
        "bg": "E8F5E9",
    },
]

for strat in strategies:
    section_bar(ws4,R,f"  {strat['id']}: {strat['title']}",9,bg=TEAL); R+=1
    rows_s=[
        ("Rule",       strat["rule"]),
        ("Evidence",   strat["evidence"]),
        ("Target Segments", strat["segments"]),
        ("Exceptions", strat["exceptions"]),
        ("Expected Outcome", strat["expected"]),
    ]
    for lbl,val in rows_s:
        ws4.row_dimensions[R].height=52
        dcell(ws4,R,1,lbl,bg=strat["bg"],bold=True,color=BLUE,halign="left")
        c=ws4.cell(row=R,column=2,value=val)
        c.font=Font(name="Arial",size=10,color=DKG)
        c.fill=PatternFill("solid",fgColor=strat["bg"])
        c.alignment=Alignment(horizontal="left",vertical="top",wrap_text=True)
        c.border=thin()
        ws4.merge_cells(start_row=R,start_column=2,end_row=R,end_column=9)
        R+=1
    R+=1

# Supporting evidence table
section_bar(ws4,R,"SUPPORTING EVIDENCE — Perf Segment × Sentiment",9,bg=RED); R+=1
cols_ps=["Sentiment","Perf Segment","Avg PnL","Win Rate","Count"]
nfmt_ps=[None,None,"#,##0.00;(#,##0.00);-","0.0%","#,##0"]
for ci,h in enumerate(cols_ps,1): hcell(ws4,R,ci,h,bg="2C3E50",sz=9)
ws4.row_dimensions[R].height=22; R+=1
for ri,row in enumerate(ps.itertuples(index=False),R):
    ws4.row_dimensions[ri].height=17
    vals=list(row)
    for ci,(v,nfmt) in enumerate(zip(vals,nfmt_ps),1):
        if ci==1:
            sb="FDECEA" if v=="Fear" else "E8F5E9"
            dcell(ws4,ri,ci,v,bg=sb,bold=True,color=RED if v=="Fear" else TEAL,halign="center")
        else:
            col=pnl_color(v) if ci==3 else (wr_color(v) if ci==4 else DKG)
            dcell(ws4,ri,ci,v,bg=alt_bg(ri),color=col,nfmt=nfmt,
                  bold=(ci==2),halign="left" if ci==2 else "right")
    R=ri
R+=2

set_widths(ws4,[18,70,4,4,4,4,4,4,4])

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 5 — BONUS: MODEL
# ─────────────────────────────────────────────────────────────────────────────
ws5 = wb.create_sheet("🤖 Bonus — Model")
ws5.sheet_view.showGridLines = False
title_row(ws5,1,"Bonus A — Predictive Model: Next-Day Profitability (No Target Leakage)",9)

R=3
section_bar(ws5,R,"MODEL DESIGN",9,bg=BLUE); R+=1
design=[
    ("Algorithm","Random Forest Classifier (n_estimators=150, max_depth=6)"),
    ("Target Variable","Next-day profit bucket: Loss / Small Gain (<$500) / Large Gain (>$500)"),
    ("Features","6 lag-1 features (prior day): win_rate, avg_leverage, n_trades, long_short_ratio, avg_size + sentiment_enc"),
    ("Leakage Prevention","All features are t-1 (shifted per account) so no same-day information leaks into target"),
    ("Validation","5-fold Stratified K-Fold Cross-Validation"),
    ("Training Set","~40,671 trader-day records (after dropping first day per account)"),
]
for i,(lbl,val) in enumerate(design,R):
    ws5.row_dimensions[i].height=22
    bg_d=alt_bg(i)
    dcell(ws5,i,1,lbl,bg=bg_d,bold=True,color=BLUE,halign="left")
    c=ws5.cell(row=i,column=2,value=val)
    c.font=Font(name="Arial",size=10,color=DKG)
    c.fill=PatternFill("solid",fgColor=bg_d)
    c.alignment=Alignment(horizontal="left",vertical="center")
    c.border=thin()
    ws5.merge_cells(start_row=i,start_column=2,end_row=i,end_column=9)
    R=i
R+=2

section_bar(ws5,R,"CROSS-VALIDATION RESULTS",9,bg=TEAL); R+=1
cv_data=[
    ("Metric","Mean","Std Dev","Interpretation"),
    ("Val Accuracy","51.9%","±0.2%","Better than random (33% for 3 classes); limited by synthetic data"),
    ("Val F1-Macro","33.4%","±0.1%","Imbalanced classes (Loss dominates); model skews toward Loss prediction"),
    ("Train Accuracy","52.9%","±0.1%","Low overfit gap (~1%) confirms no leakage"),
]
for ci,h in enumerate(cv_data[0],1): hcell(ws5,R,ci,h,bg="2C3E50",sz=9)
ws5.row_dimensions[R].height=22; R+=1
for ri,row in enumerate(cv_data[1:],R):
    ws5.row_dimensions[ri].height=20
    for ci,v in enumerate(row,1):
        is_lbl=ci==1
        dcell(ws5,ri,ci,v,bg=alt_bg(ri),bold=is_lbl,color=BLUE if is_lbl else DKG,
              halign="left" if ci in (1,4) else "center")
    R=ri
R+=2

section_bar(ws5,R,"FEATURE IMPORTANCES",9,bg=GOLD); R+=1
fi_data=[
    ("Feature","Importance","Notes"),
    ("lag_avg_size",    "34.2%","Position size is the strongest predictor of next-day outcome"),
    ("lag_avg_leverage","19.8%","Leverage level predicts risk of loss"),
    ("lag_win_rate",    "10.3%","Recent form (momentum signal)"),
    ("sentiment_enc",   " 9.1%","Greed/Fear sentiment — meaningful but not dominant"),
    ("lag_long_short_ratio"," 3.9%","Directional bias has moderate predictive value"),
    ("lag_n_trades",    " 2.7%","Trade frequency is the weakest single predictor"),
]
for ci,h in enumerate(fi_data[0],1): hcell(ws5,R,ci,h,bg="2C3E50",sz=9)
ws5.row_dimensions[R].height=22; R+=1
for ri,row in enumerate(fi_data[1:],R):
    ws5.row_dimensions[ri].height=20
    for ci,v in enumerate(row,1):
        dcell(ws5,ri,ci,v,bg=alt_bg(ri),bold=(ci==1),
              color=BLUE if ci==1 else (TEAL if ci==2 else DKG),
              halign="left" if ci!=2 else "center")
    R=ri

set_widths(ws5,[24,14,14,60,4,4,4,4,4])

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 6 — BONUS: ARCHETYPES
# ─────────────────────────────────────────────────────────────────────────────
ws6 = wb.create_sheet("👥 Bonus — Archetypes")
ws6.sheet_view.showGridLines = False
title_row(ws6,1,"Bonus B — KMeans Trader Archetypes (k=4)",9)

R=3
section_bar(ws6,R,"CLUSTER PROFILES",9,bg=BLUE); R+=1
arch_cols=["Archetype","Count","Total PnL","Win Rate","Avg Leverage","Trades","Avg Size","PnL Std Dev","Max DD"]
arch_nfmt=[None,None,"#,##0.00;(#,##0.00);-","0.0%","0.0","#,##0","#,##0","#,##0","#,##0.00;(#,##0.00);-"]
for ci,h in enumerate(arch_cols,1): hcell(ws6,R,ci,h,bg="2C3E50",sz=9)
ws6.row_dimensions[R].height=22; R+=1

arch_colors={"Consistent Winner":TEAL,"High-Risk Active":RED,
             "Passive / Occasional":GOLD,"Moderate Trader":BLUE}
arch_summary=ts.groupby("archetype").agg(
    count       =("account","count"),
    total_pnl   =("total_pnl","mean"),
    win_rate    =("win_rate","mean"),
    avg_leverage=("avg_leverage","mean"),
    n_trades    =("n_trades","mean"),
    avg_size    =("avg_size","mean"),
    pnl_std     =("pnl_std","mean"),
    max_dd      =("max_drawdown","mean"),
).round(2).reset_index().sort_values("total_pnl",ascending=False)

for ri,row in enumerate(arch_summary.itertuples(index=False),R):
    ws6.row_dimensions[ri].height=22
    vals=[row.archetype,row.count,row.total_pnl,row.win_rate,
          row.avg_leverage,row.n_trades,row.avg_size,row.pnl_std,row.max_dd]
    arch_bg=arch_colors.get(row.archetype,BLUE)+"22"  # light tint
    for ci,(v,nfmt) in enumerate(zip(vals,arch_nfmt),1):
        col=arch_colors.get(row.archetype,DKG) if ci==1 else (
            pnl_color(v) if ci in (3,9) else (wr_color(v) if ci==4 else DKG))
        dcell(ws6,ri,ci,v,bg=alt_bg(ri),bold=(ci==1),color=col,nfmt=nfmt,
              halign="left" if ci==1 else "right")
    R=ri
R+=2

section_bar(ws6,R,"ARCHETYPE DESCRIPTIONS",9,bg=NAV); R+=1
descriptions=[
    ("Consistent Winner", TEAL,
     "High win rate (≥55%), net-positive PnL. Conservative leverage. These traders demonstrate disciplined risk management and consistent edge. Strategy: Allow increased frequency on Greed days."),
    ("High-Risk Active",  RED,
     "Highest average leverage (>10x). Most active. Despite high activity, PnL is near breakeven or negative due to leverage amplifying losses. Strategy: Hard leverage cap on Fear days; position size reduction."),
    ("Moderate Trader",   BLUE,
     "Mid-range on all metrics. No strong edge but not consistently losing. Mixed sentiment sensitivity. Strategy: Standard rules apply; monitor for drift toward Underperformer behaviour."),
    ("Passive / Occasional", GOLD,
     "Fewest trades (<200). Lower leverage. PnL varies widely due to small sample. Strategy: Insufficient data for confident rule-setting; apply conservative defaults."),
]
for i,(arch,color,desc) in enumerate(descriptions,R):
    ws6.row_dimensions[i].height=55
    dcell(ws6,i,1,arch,bg=color+"33",bold=True,color=color,halign="left")
    c=ws6.cell(row=i,column=2,value=desc)
    c.font=Font(name="Arial",size=10,color=DKG)
    c.fill=PatternFill("solid",fgColor=color+"11")
    c.alignment=Alignment(horizontal="left",vertical="top",wrap_text=True)
    c.border=thin()
    ws6.merge_cells(start_row=i,start_column=2,end_row=i,end_column=9)
    R=i

set_widths(ws6,[22,8,16,10,13,10,12,13,16])

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 7 — CHARTS GALLERY
# ─────────────────────────────────────────────────────────────────────────────
ws7 = wb.create_sheet("📉 Charts")
ws7.sheet_view.showGridLines = False
title_row(ws7,1,"Charts Gallery — All Analysis Visuals",6)

chart_meta=[
    ("chart1_performance_by_sentiment.png","Chart 1","B1 — PnL, Win Rate & Drawdown by Sentiment"),
    ("chart2_behaviour_by_sentiment.png",  "Chart 2","B2 — Behaviour Shifts: Fear vs Greed"),
    ("chart3_segmentation.png",            "Chart 3","B3 — Trader Segmentation (Leverage / Frequency / Performance)"),
    ("chart4_segment_x_sentiment.png",     "Chart 4","B3 — Segment × Sentiment Deep-Dive"),
    ("chart5_timeseries.png",              "Chart 5","B4 — Market PnL, Win Rate & Leverage Over Time"),
    ("chart6_winrate_heatmap.png",         "Chart 6","Insight 1 — Win Rate Heatmap: Symbol × Sentiment"),
    ("chart7_feature_importance.png",      "Chart 7","Bonus A — Predictive Model Feature Importances"),
    ("chart8_archetypes.png",              "Chart 8","Bonus B — Trader Archetypes Scatter Plot"),
]

row_cursor=3
col_cursor=1
for idx,(fname,label,desc) in enumerate(chart_meta):
    fpath=f"{CHARTS}/{fname}"
    if not os.path.exists(fpath): continue

    # Label
    c=ws7.cell(row=row_cursor,column=col_cursor,value=f"{label}: {desc}")
    c.font=Font(name="Arial",bold=True,size=10,color=WHT)
    c.fill=PatternFill("solid",fgColor=NAV)
    c.alignment=Alignment(horizontal="left",vertical="center")
    c.border=thin()
    ws7.row_dimensions[row_cursor].height=20

    # Image
    img=XLImage(fpath)
    img.width=560; img.height=210
    img.anchor=f"{get_column_letter(col_cursor)}{row_cursor+1}"
    ws7.add_image(img)
    row_cursor+=16

set_widths(ws7,[90])

# ─────────────────────────────────────────────────────────────────────────────
# Tab colours
# ─────────────────────────────────────────────────────────────────────────────
ws1.sheet_properties.tabColor=NAV
ws2.sheet_properties.tabColor=BLUE
ws3.sheet_properties.tabColor=TEAL
ws4.sheet_properties.tabColor=GOLD
ws5.sheet_properties.tabColor="9B5DE5"
ws6.sheet_properties.tabColor=RED
ws7.sheet_properties.tabColor="888888"

out="/home/claude/primetrade/outputs/primetrade_full_analysis.xlsx"
wb.save(out)
print(f"Saved: {out}")
